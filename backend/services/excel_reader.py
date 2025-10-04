"""Excel file reading and sheet management services."""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from backend.core.logging import logger


class ExcelReader:
    """Service for reading and managing Excel workbooks."""

    def __init__(self, file_path: Path):
        """Initialize with Excel file path."""
        self.file_path = file_path
        self.workbook = None
        self._sheet_names = None

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names from the workbook."""
        if self._sheet_names is None:
            try:
                self.workbook = load_workbook(self.file_path, read_only=True)
                self._sheet_names = self.workbook.sheetnames
                logger.info(
                    f"Found {len(self._sheet_names)} sheets",
                    file=str(self.file_path),
                    sheets=self._sheet_names,
                )
            except Exception as e:
                logger.error(f"Failed to read workbook: {e}", file=str(self.file_path))
                raise ValueError(f"Invalid Excel file: {e}")

        return self._sheet_names

    def read_sheet(
        self,
        sheet_name: str,
        dtype: str = "str",
        nrows: Optional[int] = None,
        clean_headers: bool = True,
    ) -> pd.DataFrame:
        """Read a specific sheet as DataFrame."""
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                dtype=dtype,
                nrows=nrows,
                engine="openpyxl",
            )

            # Clean multi-row headers if requested
            if clean_headers:
                df = self._clean_multi_row_headers(df, sheet_name)

            logger.info(
                f"Read sheet '{sheet_name}'", rows=len(df), cols=len(df.columns)
            )

            return df

        except Exception as e:
            logger.error(f"Failed to read sheet '{sheet_name}': {e}")
            raise ValueError(f"Could not read sheet '{sheet_name}': {e}")

    def _clean_multi_row_headers(
        self, df: pd.DataFrame, sheet_name: str
    ) -> pd.DataFrame:
        """Clean multi-row headers that are common in Excel files."""
        if len(df) == 0:
            return df

        # Look for rows that appear to be continuation of headers
        rows_to_remove = []

        for i in range(min(5, len(df))):  # Check first 5 rows
            row = df.iloc[i]
            row_values = [
                str(val).strip()
                for val in row.values
                if pd.notna(val) and str(val).strip()
            ]

            if not row_values:  # Empty row
                continue

            # Check if this row looks like header continuation
            # Indicators: contains parentheses, "remarks", "status", "date", etc.
            header_indicators = [
                "(",
                ")",
                "remarks",
                "status",
                "date",
                "details",
                "deviation",
                "under",
                "available",
                "promised",
                "ok/nok",
                "yes/no",
            ]

            indicator_count = sum(
                1
                for val in row_values
                for indicator in header_indicators
                if indicator.lower() in val.lower()
            )

            # If more than 30% of non-empty values contain header indicators
            if len(row_values) > 0 and indicator_count / len(row_values) > 0.3:
                # Additional check: make sure it's not actual data
                # Real data usually has part numbers (numeric) or specific patterns
                numeric_like = sum(
                    1
                    for val in row_values
                    if val.replace("-", "").replace(".", "").isdigit() and len(val) > 3
                )

                # If less than 20% looks like part numbers, likely a header row
                if numeric_like / len(row_values) < 0.2:
                    rows_to_remove.append(i)
                    logger.info(
                        f"Detected header continuation row {i} in {sheet_name}: {row_values[:3]}"
                    )

        # Remove detected header rows
        if rows_to_remove:
            df = df.drop(index=rows_to_remove).reset_index(drop=True)
            logger.info(
                f"Removed {len(rows_to_remove)} header continuation rows from {sheet_name}"
            )

        return df

    def get_sheet_info(self, sheet_name: str) -> Dict[str, any]:
        """Get basic information about a sheet."""
        df = self.read_sheet(
            sheet_name, nrows=1, clean_headers=False
        )  # Just read header

        # Get full sheet to count rows (inefficient but simple)
        full_df = self.read_sheet(sheet_name, clean_headers=False)

        return {
            "name": sheet_name,
            "total_rows": len(full_df),
            "total_cols": len(df.columns),
            "columns": df.columns.tolist(),
        }

    def preview_sheet(self, sheet_name: str, n: int = 10) -> Dict[str, any]:
        """Get preview of sheet with head and tail samples."""
        df = self.read_sheet(sheet_name)

        # Clean column names for JSON serialization
        df.columns = [str(col).strip() for col in df.columns]

        # Get head and tail samples
        head_data = df.head(n).fillna("").to_dict("records")
        tail_data = df.tail(n).fillna("").to_dict("records")

        return {
            "sheet_name": sheet_name,
            "total_rows": len(df),
            "total_cols": len(df.columns),
            "columns": df.columns.tolist(),
            "head_data": head_data,
            "tail_data": tail_data,
        }

    def detect_project_columns(
        self, df: pd.DataFrame, id_col: str = "YAZAKI PN"
    ) -> Tuple[List[str], int, int]:
        """
        Detect project/plant columns in MasterBOM sheet.

        Returns:
            - List of project column names
            - Start index of project columns
            - End index of project columns
        """
        columns = df.columns.tolist()

        try:
            # Find ID column index
            id_col_idx = None
            for i, col in enumerate(columns):
                if str(col).strip().upper() == id_col.upper():
                    id_col_idx = i
                    break

            if id_col_idx is None:
                logger.warning(f"ID column '{id_col}' not found, using first column")
                id_col_idx = 0

            # Find "Item Description" or similar column
            desc_col_idx = None
            desc_patterns = [
                r"item\s*desc",
                r"description",
                r"part\s*desc",
                r"component\s*desc",
            ]

            for i, col in enumerate(columns[id_col_idx + 1 :], start=id_col_idx + 1):
                col_clean = str(col).strip().lower()
                for pattern in desc_patterns:
                    if re.search(pattern, col_clean):
                        desc_col_idx = i
                        break
                if desc_col_idx:
                    break

            if desc_col_idx is None:
                logger.warning(
                    "Description column not found, assuming all columns after ID are projects"
                )
                desc_col_idx = len(columns)

            # Project columns are between ID and Description
            project_start = id_col_idx + 1
            project_end = desc_col_idx

            project_columns = columns[project_start:project_end]

            logger.info(
                f"Detected {len(project_columns)} project columns",
                start_idx=project_start,
                end_idx=project_end,
                columns=project_columns[:5],
            )  # Log first 5

            return project_columns, project_start, project_end

        except Exception as e:
            logger.error(f"Failed to detect project columns: {e}")
            return [], 1, 1

    def close(self):
        """Close the workbook if open."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
