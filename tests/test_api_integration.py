"""Integration tests for the complete API workflow."""

import asyncio

import pandas as pd
import pytest
from httpx import AsyncClient
from openpyxl import Workbook

from backend.main import app


@pytest.mark.integration
@pytest.mark.asyncio
class TestAPIIntegration:
    """Integration tests for the complete ETL workflow."""

    @pytest.fixture
    async def client(self):
        """Create test client."""
        async with AsyncClient(app=app, base_url="http://test") as ac:
            yield ac

    @pytest.fixture
    def sample_excel_file(self, temp_dir):
        """Create a sample Excel file for testing."""
        file_path = temp_dir / "test_workbook.xlsx"

        # Create workbook with sample data
        wb = Workbook()

        # MasterBOM sheet
        ws_master = wb.active
        ws_master.title = "MasterBOM"

        # Headers
        headers = [
            "YAZAKI PN",
            "BX726_BEV_JOB1+90_YMOK",
            "BX726_BEV_JOB1+90_YRL",
            "BX726_MCA_JOB1+90_YMOK",
            "Item Description",
            "Supplier Name",
            "PSW",
            "Approved Date",
            "FAR Status",
        ]

        for col, header in enumerate(headers, 1):
            ws_master.cell(row=1, column=col, value=header)

        # Sample data
        data = [
            [
                "7009-6933",
                "X",
                "",
                "D",
                "Wire Harness",
                "Yazaki Corp",
                "OK",
                "2024-01-15",
                "OK",
            ],
            [
                "7116-4101-02",
                "D",
                "X",
                "",
                "Connector",
                "Molex Inc",
                "NG",
                "2024-02-20",
                "NG",
            ],
            [
                "ABC-123",
                "",
                "0",
                "X",
                "Terminal",
                "TE Connectivity",
                "",
                "2024-03-25",
                "Pending",
            ],
        ]

        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws_master.cell(row=row, column=col, value=value)

        # Status sheet
        ws_status = wb.create_sheet("Status")

        status_headers = ["OEM", "Project", "Total Part Numbers", "PSW Available"]
        for col, header in enumerate(status_headers, 1):
            ws_status.cell(row=1, column=col, value=header)

        status_data = [
            ["Toyota", "BX726 BEV", 1250, "85%"],
            ["Honda", "HX123 ICE", 890, "92%"],
        ]

        for row, row_data in enumerate(status_data, 2):
            for col, value in enumerate(row_data, 1):
                ws_status.cell(row=row, column=col, value=value)

        wb.save(file_path)
        return file_path

    async def test_complete_workflow(self, client, sample_excel_file):
        """Test the complete ETL workflow from upload to results."""

        # Step 1: Upload file
        with open(sample_excel_file, "rb") as f:
            files = {
                "file": (
                    "test_workbook.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = await client.post("/api/upload", files=files)

        assert upload_response.status_code == 200
        upload_data = upload_response.json()

        assert "file_id" in upload_data
        assert "sheet_names" in upload_data
        assert "MasterBOM" in upload_data["sheet_names"]
        assert "Status" in upload_data["sheet_names"]

        file_id = upload_data["file_id"]

        # Step 2: Preview sheets
        preview_response = await client.get(
            f"/api/preview?file_id={file_id}&sheet=MasterBOM&n=5"
        )

        assert preview_response.status_code == 200
        preview_data = preview_response.json()

        assert preview_data["sheet_name"] == "MasterBOM"
        assert preview_data["total_rows"] == 3
        assert len(preview_data["head_data"]) <= 5
        assert "YAZAKI PN" in preview_data["columns"]

        # Step 3: Profile sheets
        profile_response = await client.get(
            f"/api/profile?file_id={file_id}&sheet=MasterBOM"
        )

        assert profile_response.status_code == 200
        profile_data = profile_response.json()

        assert profile_data["sheet_name"] == "MasterBOM"
        assert profile_data["total_rows"] == 3
        assert len(profile_data["columns"]) > 0

        # Step 4: Transform data
        transform_request = {
            "file_id": file_id,
            "master_sheet": "MasterBOM",
            "status_sheet": "Status",
            "options": {"id_col": "YAZAKI PN", "date_cols": ["Approved Date"]},
        }

        transform_response = await client.post("/api/transform", json=transform_request)

        assert transform_response.status_code == 200
        transform_data = transform_response.json()

        assert transform_data["success"] is True
        assert "artifacts" in transform_data
        assert "summary" in transform_data
        assert len(transform_data["artifacts"]) > 0

        # Verify summary
        summary = transform_data["summary"]
        assert summary["total_parts"] > 0
        assert summary["processing_time_seconds"] > 0

        # Verify artifacts
        artifacts = transform_data["artifacts"]
        artifact_names = [artifact["name"] for artifact in artifacts]

        expected_artifacts = [
            "masterbom_clean.csv",
            "masterbom_clean.parquet",
            "plant_item_status.csv",
            "plant_item_status.parquet",
            "fact_parts.csv",
            "fact_parts.parquet",
            "status_clean.csv",
            "status_clean.parquet",
            "dim_dates.csv",
            "dim_dates.parquet",
            "etl.sqlite",
        ]

        for expected in expected_artifacts:
            assert expected in artifact_names

    async def test_error_handling(self, client):
        """Test error handling in the API."""

        # Test invalid file upload
        files = {"file": ("test.txt", b"not an excel file", "text/plain")}
        upload_response = await client.post("/api/upload", files=files)

        assert upload_response.status_code == 400

        # Test preview with invalid file ID
        preview_response = await client.get(
            "/api/preview?file_id=invalid-id&sheet=Sheet1"
        )

        assert preview_response.status_code == 400

        # Test transform with missing file
        transform_request = {
            "file_id": "non-existent-file",
            "master_sheet": "MasterBOM",
            "status_sheet": "Status",
            "options": {},
        }

        transform_response = await client.post("/api/transform", json=transform_request)

        assert transform_response.status_code == 404

    async def test_file_cleanup(self, client, sample_excel_file):
        """Test file cleanup functionality."""

        # Upload file
        with open(sample_excel_file, "rb") as f:
            files = {
                "file": (
                    "test_workbook.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = await client.post("/api/upload", files=files)

        assert upload_response.status_code == 200
        file_id = upload_response.json()["file_id"]

        # Delete file
        delete_response = await client.delete(f"/api/upload/{file_id}")

        assert delete_response.status_code == 200

        # Verify file is deleted - preview should fail
        preview_response = await client.get(
            f"/api/preview?file_id={file_id}&sheet=MasterBOM"
        )

        assert preview_response.status_code == 404

    async def test_large_file_handling(self, client, temp_dir):
        """Test handling of larger files."""

        # Create larger Excel file
        file_path = temp_dir / "large_workbook.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "MasterBOM"

        # Headers
        headers = [
            "YAZAKI PN",
            "PROJECT_1",
            "PROJECT_2",
            "Item Description",
            "Supplier Name",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Generate more data (100 rows)
        for row in range(2, 102):
            ws.cell(row=row, column=1, value=f"PART-{row:04d}")
            ws.cell(row=row, column=2, value="X" if row % 2 == 0 else "")
            ws.cell(row=row, column=3, value="D" if row % 3 == 0 else "")
            ws.cell(row=row, column=4, value=f"Component {row}")
            ws.cell(row=row, column=5, value=f"Supplier {row % 5}")

        # Add Status sheet
        ws_status = wb.create_sheet("Status")
        ws_status.cell(row=1, column=1, value="OEM")
        ws_status.cell(row=1, column=2, value="Project")
        ws_status.cell(row=2, column=1, value="Toyota")
        ws_status.cell(row=2, column=2, value="Large Project")

        wb.save(file_path)

        # Upload and process
        with open(file_path, "rb") as f:
            files = {
                "file": (
                    "large_workbook.xlsx",
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            upload_response = await client.post("/api/upload", files=files)

        assert upload_response.status_code == 200
        file_id = upload_response.json()["file_id"]

        # Transform
        transform_request = {
            "file_id": file_id,
            "master_sheet": "MasterBOM",
            "status_sheet": "Status",
            "options": {"id_col": "YAZAKI PN"},
        }

        transform_response = await client.post("/api/transform", json=transform_request)

        assert transform_response.status_code == 200
        transform_data = transform_response.json()

        assert transform_data["success"] is True
        assert transform_data["summary"]["total_parts"] == 100

    async def test_concurrent_requests(self, client, sample_excel_file):
        """Test handling of concurrent requests."""

        async def upload_and_process():
            with open(sample_excel_file, "rb") as f:
                files = {
                    "file": (
                        "test_workbook.xlsx",
                        f,
                        (
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                    )
                }
                upload_response = await client.post("/api/upload", files=files)

            if upload_response.status_code != 200:
                return False

            file_id = upload_response.json()["file_id"]

            transform_request = {
                "file_id": file_id,
                "master_sheet": "MasterBOM",
                "status_sheet": "Status",
                "options": {"id_col": "YAZAKI PN"},
            }

            transform_response = await client.post(
                "/api/transform", json=transform_request
            )
            return transform_response.status_code == 200

        # Run multiple concurrent requests
        tasks = [upload_and_process() for _ in range(3)]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # At least some should succeed
        successful = sum(1 for result in results if result is True)
        assert successful > 0

    async def test_health_endpoints(self, client):
        """Test health check endpoints."""

        # Root endpoint
        root_response = await client.get("/")
        assert root_response.status_code == 200

        root_data = root_response.json()
        assert root_data["status"] == "healthy"

        # Health endpoint
        health_response = await client.get("/health")
        assert health_response.status_code == 200

        health_data = health_response.json()
        assert health_data["status"] == "healthy"


@pytest.mark.integration
class TestDataValidation:
    """Test data validation throughout the pipeline."""

    def test_id_cleaning_consistency(self, sample_excel_data):
        """Test that ID cleaning is consistent across the pipeline."""
        from backend.core.logging import ETLLogger
        from backend.services.cleaning import clean_id
        from backend.services.masterbom_rules import MasterBOMProcessor

        # Test individual function
        test_ids = ["  7009-6933  ", "7116@4101#02", "ABC_123"]
        cleaned_individual = [clean_id(id_val) for id_val in test_ids]

        # Test through processor
        df = pd.DataFrame({"YAZAKI PN": test_ids, "Item Description": ["A", "B", "C"]})
        processor = MasterBOMProcessor(df, ETLLogger())
        processor.id_column = "YAZAKI PN"
        processor._clean_id_column()

        cleaned_processor = processor.df["part_id_std"].tolist()

        # Should be identical
        assert cleaned_individual == cleaned_processor

    def test_date_parsing_consistency(self):
        """Test that date parsing is consistent."""
        from backend.services.cleaning import parse_date_column

        test_dates = pd.Series(["2024-01-15", "2024-02-20", "invalid", "2024-03-25"])

        result1 = parse_date_column(test_dates, "test_date")
        result2 = parse_date_column(test_dates, "test_date")

        # Results should be identical
        pd.testing.assert_frame_equal(result1, result2)

    def test_status_classification_rules(self):
        """Test that status classification rules are applied correctly."""
        from backend.core.logging import ETLLogger
        from backend.services.masterbom_rules import MasterBOMProcessor

        processor = MasterBOMProcessor(pd.DataFrame(), ETLLogger())

        # Test all status cases
        test_cases = [
            ("X", "active"),
            ("D", "inactive"),
            ("0", "duplicate"),
            ("", "new"),
            (None, "new"),
            ("unknown", "new"),
        ]

        for raw_status, expected_class in test_cases:
            row = pd.Series({"raw_status": raw_status})
            result = processor._classify_status(row)
            assert (
                result == expected_class
            ), f"Failed for {raw_status}: expected {expected_class}, got {result}"


@pytest.mark.slow
@pytest.mark.integration
class TestPerformance:
    """Performance tests for the ETL pipeline."""

    def test_large_dataset_processing(self, temp_dir):
        """Test processing of large datasets."""
        from backend.core.logging import ETLLogger
        from backend.services.masterbom_rules import MasterBOMProcessor
        from tests.conftest import generate_large_masterbom

        # Generate large dataset
        large_df = generate_large_masterbom(n_parts=5000, n_projects=20)

        processor = MasterBOMProcessor(large_df, ETLLogger())

        import time

        start_time = time.time()

        result = processor.process(id_col="YAZAKI PN")

        processing_time = time.time() - start_time

        # Should complete in reasonable time (adjust threshold as needed)
        assert processing_time < 30  # 30 seconds

        # Verify results
        assert len(result["masterbom_clean"]) == 5000
        assert len(result["fact_parts"]) == 5000
        assert len(result["plant_item_status"]) == 5000 * 20

    def test_memory_usage(self):
        """Test memory usage with large datasets."""
        import os

        import psutil

        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss

        # Process large dataset
        from backend.core.logging import ETLLogger
        from backend.services.masterbom_rules import MasterBOMProcessor
        from tests.conftest import generate_large_masterbom

        large_df = generate_large_masterbom(n_parts=10000, n_projects=10)
        processor = MasterBOMProcessor(large_df, ETLLogger())
        processor.process(id_col="YAZAKI PN")

        peak_memory = process.memory_info().rss
        memory_increase = peak_memory - initial_memory

        # Memory increase should be reasonable (adjust threshold as needed)
        # Allow up to 500MB increase for 10k parts
        assert memory_increase < 500 * 1024 * 1024  # 500MB
